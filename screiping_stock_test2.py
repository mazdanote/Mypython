import time
import os
import pandas as pd
import openpyxl
from selenium import webdriver

driver = webdriver.Chrome() # WebDriverのインスタンスを作成
driver.get('https://financials.morningstar.com/ratios/r.html?t=0P000000GY') # URLを指定してブラウザを開く
time.sleep(5) # 5秒待機

# get width and height of the page
w = driver.execute_script("return document.body.scrollWidth;")
h = driver.execute_script("return document.body.scrollHeight;")

# set window size
driver.set_window_size(w,h)

# Get Screen Shot
FILENAME = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screen2.png")
driver.save_screenshot(FILENAME)

# print(driver.page_source)
f = open('screen.html','w')
f.write(driver.page_source)
f.close()

crypto_data = pd.read_html(driver.page_source)
print(crypto_data[0])
time.sleep(10) # 10秒待機

wb=openpyxl.Workbook()
wb.create_sheet(index=0, title='APPL')
wb.create_sheet(index=1, title='CRWD')
crypto_data[0].to_excel('testtest.xlsx', sheet_name='APPL') #crypto_dataをAPPLのシートに入力する

#==========ここから追加作業
# 新しいタブを作成する
driver.execute_script("window.open()")

# 新しいタブに切り替える
driver.switch_to.window(driver.window_handles[1])

driver.get('https://financials.morningstar.com/ratios/r.html?t=0P0001HOOZ&culture=en&platform=sal') # CRWDのURLを指定してブラウザを開く
time.sleep(5) # 5秒待機

# get width and height of the page
w = driver.execute_script("return document.body.scrollWidth;")
h = driver.execute_script("return document.body.scrollHeight;")

# set window size
driver.set_window_size(w,h)

# Get Screen Shot
FILENAME = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screen3.png")
driver.save_screenshot(FILENAME)

# print(driver.page_source)
f = open('screen.html','w')
f.write(driver.page_source)
f.close()

crypto_data1 = pd.read_html(driver.page_source)
print(crypto_data1[1])
time.sleep(10) # 15秒待機

crypto_data1[1].to_excel('testtest.xlsx', sheet_name="CRWD") # crypto_data1をCRWDのシートに記載する
wb.save('testtest.xlsx')

driver.quit() # ブラウザを閉じる
#==========ここから別作業
